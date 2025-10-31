"""
json2excel.py - JSON 상품 트리 → 엑셀 변환 자동화 도구

이 스크립트는 다음과 같은 기능을 제공합니다:
- 단일 JSON 파일 또는 디렉토리 내 여러 JSON 파일에서 상품 트리 정보를 추출하여 엑셀 파일(.xlsx)로 변환합니다.
- 상품 트리의 각 Product 노드에서 상품명, 경로, 가격, 수익 등 주요 정보를 추출합니다.
- 변환된 엑셀 파일은 'excels' 하위 폴더에 sitecode-날짜시간.xlsx 형식으로 저장됩니다.
- 명령행 옵션으로 단일 파일(--file, --sitecode) 또는 폴더 전체(--dir) 변환을 지원합니다.
- 모든 실행/오류/경고 메시지는 프로젝트 표준 logger(영문, [YY/MM/DD HH:MM:SS Level json2excel] 포맷)로 출력됩니다.
- 주요 함수/클래스/코드 블록은 한글로 문서화되어 있습니다.

사용 예시:
    python json2excel.py --file data.json --sitecode UK
    python json2excel.py --dir ./jsons
"""

import argparse
import json
import os
from datetime import datetime
from typing import Any, List, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
import glob
from dataclasses import dataclass
from utility.orangelogger import log

# === 데이터 클래스 정의 ===
@dataclass
class ProductInfo:
    번호: int
    경로: str
    상품명: str
    URL: str
    ModelPrice: str
    ModelRevenue: str

# === 천단위 콤마 포맷 함수 ===
def format_number(value: str) -> str:
    """
    천단위 콤마 포맷 함수
    Args:
        value (str): 숫자 문자열
    Returns:
        str: 콤마가 적용된 문자열
    """
    if value is None or value == '':
        return ''
    try:
        return f"{int(value):,}"
    except (ValueError, TypeError):
        return ''

# === 트리 탐색 및 Product 추출 ===
def extract_products(tree: List[Dict[str, Any]], path: List[str], products: List[ProductInfo]) -> None:
    """
    트리 구조에서 상품(Product) 노드를 추출하여 products 리스트에 추가합니다.
    Args:
        tree (List[Dict[str, Any]]): 상품 트리
        path (List[str]): 현재까지의 경로
        products (List[ProductInfo]): 추출 결과 리스트
    """
    for node in tree:
        node_type = node.get('node_type', '')
        name = node.get('name', '')
        url = node.get('url', '')
        meta = node.get('meta', {})
        children = node.get('children', [])
        if node_type == 'Product':
            model_price = format_number(meta.get('data-modelprice', ''))
            model_revenue = format_number(meta.get('data-modelrevenue', ''))
            full_path = ' / '.join(path + [name])
            products.append(ProductInfo(
                번호=len(products) + 1,
                경로=full_path,
                상품명=name,
                URL=url,
                ModelPrice=model_price,
                ModelRevenue=model_revenue
            ))
        else:
            extract_products(children, path + [name], products)

def extract_sitecode_from_filename(filename: str) -> str:
    """
    파일명에서 sitecode를 추출합니다.
    예: UK_shop_250606-175811_xxx.json -> UK
    Args:
        filename (str): 파일명
    Returns:
        str: sitecode
    """
    base = os.path.basename(filename)
    if '_shop' in base:
        return base.split('_shop')[0]
    return ''

# === 메인 함수 ===
def main() -> None:
    """
    명령행 인자를 받아 JSON 상품 트리를 엑셀로 변환합니다.
    """
    parser = argparse.ArgumentParser(description='Extract product info from json and save to excel')
    parser.add_argument('--file', help='Path to json file')
    parser.add_argument('--sitecode', help='Sitecode for sheet and filename')
    parser.add_argument('--dir', help='Directory containing json files')
    args = parser.parse_args()

    excels_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excels')
    os.makedirs(excels_dir, exist_ok=True)

    if args.dir:
        # 폴더 내 모든 json 파일 처리
        json_files = glob.glob(os.path.join(args.dir, '*.json'))
        if not json_files:
            log.info('No json files found in the directory.')
            return
        for json_path in json_files:
            sitecode = extract_sitecode_from_filename(json_path)
            if not sitecode:
                log.warning(f'Skip file (invalid format): {json_path}')
                continue
            try:
                with open(json_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                tree = data.get('tree', [])
                products: List[ProductInfo] = []
                extract_products(tree, [], products)
                now = datetime.now().strftime('%y%m%d-%H%M%S')
                excel_filename = f"{sitecode}-{now}.xlsx"
                excel_path = os.path.join(excels_dir, excel_filename)
                wb = Workbook()
                ws = wb.active
                ws.title = sitecode
                columns = ['번호', '경로', '상품명', 'URL', 'ModelPrice', 'ModelRevenue']
                ws.append(columns)
                for p in products:
                    ws.append([p.번호, p.경로, p.상품명, p.URL, p.ModelPrice, p.ModelRevenue])
                header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
                center_align = Alignment(horizontal='center', vertical='center')
                right_align = Alignment(horizontal='right', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                font9 = Font(size=9)
                col_settings = {
                    2: 50,  # 경로
                    3: 50,  # 상품명
                    4: 50   # URL
                }
                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    if col[0].column in col_settings:
                        ws.column_dimensions[col_letter].width = col_settings[col[0].column]
                    else:
                        max_length = 0
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except Exception:
                                pass
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 200)
                row_count = ws.max_row
                col_count = ws.max_column
                for row in ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count):
                    for cell in row:
                        cell.border = thin_border
                        cell.font = font9
                        match cell.row:
                            case 1:
                                cell.alignment = center_align
                                cell.fill = header_fill
                            case _ if cell.column in [1, 5, 6]:
                                cell.alignment = right_align
                wb.save(excel_path)
                log.info(f"Excel file created: {excel_path}")
            except Exception as e:
                log.error(f"Error processing {json_path}: {e}")
    else:
        # 기존 단일 파일 처리
        if not args.file or not args.sitecode:
            log.error('--file and --sitecode or --dir is required.')
            return
        json_path = args.file
        sitecode = args.sitecode
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        tree = data.get('tree', [])
        products: List[ProductInfo] = []
        extract_products(tree, [], products)
        now = datetime.now().strftime('%y%m%d-%H%M%S')
        excel_filename = f"{sitecode}-{now}.xlsx"
        excel_path = os.path.join(excels_dir, excel_filename)
        wb = Workbook()
        ws = wb.active
        ws.title = sitecode
        columns = ['번호', '경로', '상품명', 'URL', 'ModelPrice', 'ModelRevenue']
        ws.append(columns)
        for p in products:
            ws.append([p.번호, p.경로, p.상품명, p.URL, p.ModelPrice, p.ModelRevenue])
        header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        font9 = Font(size=9)
        col_settings = {
            2: 50,  # 경로
            3: 50,  # 상품명
            4: 50   # URL
        }
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].column in col_settings:
                ws.column_dimensions[col_letter].width = col_settings[col[0].column]
            else:
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 200)
        row_count = ws.max_row
        col_count = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count):
            for cell in row:
                cell.border = thin_border
                cell.font = font9
                match cell.row:
                    case 1:
                        cell.alignment = center_align
                        cell.fill = header_fill
                    case _ if cell.column in [1, 5, 6]:
                        cell.alignment = right_align
        wb.save(excel_path)
        log.info(f"Excel file created: {excel_path}")

if __name__ == '__main__':
    main() 