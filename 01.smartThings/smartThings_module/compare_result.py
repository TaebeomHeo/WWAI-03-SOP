import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

pd.set_option('display.max_columns', None)

class CompareProcess:
    """
    포맷 데이터와 실제 추출 데이터를 비교하여 결과를 생성하는 클래스
    
    - 포맷 데이터와 실제 데이터의 일치 여부를 비교
    - 비교 결과를 Excel 파일로 저장하고 포맷팅
    - 추천 제품 정보를 추가하여 최종 결과 생성
    """
    
    def __init__(self, df_format_data, df_abs_data,df_compare_item_path, output_path,country_code):
        """
        CompareProcess 클래스 초기화
        
        Args:
            df_format_data: 포맷 데이터 DataFrame (예상 데이터)
            df_abs_data: 실제 추출 데이터 DataFrame
            df_compare_item_path: 비교 항목 Excel 파일 경로
            output_path: 결과 Excel 파일 저장 경로
            country_code: 처리할 국가 코드 리스트
        """
        self.df_abs_data = df_abs_data  # 실제 추출된 데이터
        self.df_format_data = df_format_data  # 포맷 데이터 (예상 데이터)
        self.df_compare_item =  pd.read_excel(df_compare_item_path, header=None)  # 비교 항목 파일 (헤더 없음)
        self.country_code = country_code  # 국가 코드 리스트

        
        # 비교 결과를 저장할 DataFrame 초기화
        self.compare_result = pd.DataFrame(columns=['계정', '항목','결과','결과 상세', 'NA', '국가'])  #비교 결과를 가져오기 위한 컬럼 포맷 세팅
        self.merge_compare_result= pd.DataFrame()  # 최종 병합된 비교 결과
        self.output_path = output_path  # 출력 파일 경로
        
        # 비교할 컬럼 리스트 정의
        self.compare_columns = [ 
                        'main_headline', 'main_description',         # 메인 헤드라인, 설명
                       'storyIdRank1', 'storyIdRank2', 'storyIdRank3',  # 스토리 ID들
                       'storyIdRank1_title', 'storyIdRank1_desc',  # 스토리 1 제목/설명
                       'storyIdRank2_title', 'storyIdRank2_desc',  # 스토리 2 제목/설명
                       'storyIdRank3_title', 'storyIdRank3_desc',  # 스토리 3 제목/설명
                       'banner_text', 'banner_link_text', 'banner_hyperlink'  # 배너 관련 정보
                       ]  

    #추출한 데이터를 기준으로 비교결과에 대한 포맷 구성
    def compare_data(self):
        """
        포맷 데이터와 실제 추출 데이터를 비교하는 함수
        
        - 각 행의 각 컬럼에 대해 일치 여부를 확인
        - 공백을 제거한 상태에서 비교 수행
        - 배너 관련 컬럼은 "없음"인 경우 제외
        - 일치/불일치 결과를 compare_result DataFrame에 저장
        """
        for idx in range(len(self.df_abs_data)):

            #공백 제거한 상태의 비교 포맷 구성
            format_row = self.df_format_data.loc[idx, self.compare_columns].astype(str).apply(lambda x: ''.join(x.split())) #포맷데이터를 공백없는 상태로 변경
            abs_row = self.df_abs_data.loc[idx, self.compare_columns].astype(str).apply(lambda x: ''.join(x.split())) #추출데이터를 공백없는 상태로 변경

            # 각 비교 컬럼에 대해 일치 여부 확인
            for col in self.compare_columns:
                # 배너 관련 컬럼이 "없음"인 경우 건너뛰기 (배너는 선택적 표시)
                if abs_row[col] =="없음" and col == 'banner_text':
                    continue
                elif abs_row[col] =="없음" and col == 'banner_link_text':
                    continue
                elif abs_row[col] =="없음" and col == 'banner_hyperlink':
                    continue
                
                if col not in format_row.index: # 해당 컬럼이 존재 하지 않는 경우 스킵
                    continue
                
                if format_row[col] == abs_row[col]: # 데이터 비교 - 일치하는 경우
                    self.compare_result.loc[len(self.compare_result)] = [self.df_abs_data.loc[idx, 'Account'], col, '일치', '', '',self.df_format_data.at[idx, 'country_code']] # self.compare_result 데이터프레임에 일치 데이터 삽입
                else:  # 불일치하는 경우
                    detail = f"포맷데이터: {self.df_format_data.loc[idx, col]}\n\n\n추출데이터: {self.df_abs_data.loc[idx, col]}"  # 상세 비교 정보 생성
                    self.compare_result.loc[len(self.compare_result)] = [self.df_abs_data.loc[idx, 'Account'], col, '불일치', detail, '',self.df_format_data.at[idx, 'country_code']]# self.compare_result 데이터프레임에 불일치 데이터 삽입

        
        
    def item_abs_data(self):
        """
        비교 항목 파일을 기반으로 데이터를 필터링하는 함수
        
        - 비교 항목 파일에서 계정별 필터링 조건을 읽어옴
        - 계정이 비교 항목에 없는 경우는 모두 유지
        - 계정이 비교 항목에 있는 경우는 해당 항목만 유지
        - 최종 결과를 merge_compare_result에 저장
        """
        # 비교 항목 파일에서 계정-항목 쌍을 추출
        account_value_pairs = []
        for i in range(len(self.df_compare_item)):
            account = self.df_compare_item.iloc[i, 0]  # 첫 번째 컬럼은 계정
            values = self.df_compare_item.iloc[i, 1:].dropna().tolist()  # 나머지 컬럼들은 항목들 (NaN 제거)
            for val in values:
                account_value_pairs.append((account, val))  # 계정-항목 쌍 생성

        # 계정-항목 쌍을 DataFrame으로 변환
        account_value_df = pd.DataFrame(account_value_pairs, columns=["계정", "항목"])
        
        # 1. 계정 기준 일치 여부 확인
        accounts_in_value_df = set(account_value_df["계정"].unique())  # 비교 항목에 있는 계정들

        # 2. 계정 기준 분리
        # (a) 계정이 account_value_df에 **없는 경우** → 무조건 유지
        df_account_not_exist = self.compare_result[~self.compare_result["계정"].isin(accounts_in_value_df)]

        # (b) 계정이 존재하는 경우만 추출
        df_account_exist = self.compare_result[self.compare_result["계정"].isin(accounts_in_value_df)]

        # 계정과 항목이 모두 일치하는 데이터만 추출
        df_matched = pd.merge(df_account_exist,account_value_df,on=["계정", "항목"],how="inner")
        
        # 4. 두 결과 합치기 - 필터링된 데이터와 유지할 데이터를 병합
        self.merge_compare_result = pd.concat([df_matched, df_account_not_exist], ignore_index=True)

    #추출 데이터에서 각 계정에 맞게 추천제품 comapre_result 변수에 삽입
    #### 로직 수정해야함
    def abs_rec_data(self):
        """
        추출 데이터에서 추천 제품 정보를 비교 결과에 추가하는 함수
        
        - rec가 포함된 컬럼들을 찾아서 추천 제품 정보 수집
        - 각 계정별로 추천 제품 정보를 NA 컬럼에 추가
        - 국가별로 첫 번째 매칭되는 행에 추천 제품 정보 삽입
        """
        rec_columns = [col for col in self.df_abs_data.columns if 'rec' in col] #추출 데이터에서  rec 포함된 컬럼 추출
        
        for idx, row in self.df_abs_data.iterrows():
            country_value = row['country_code'] # 각 계정의 국가 코드 추출

            rec_product = '\n'.join(f"{col}: {row[col]}" for col in rec_columns) # 추천제품 리스트를 문자열로 결합
            match_idx = self.merge_compare_result[(self.merge_compare_result['국가'] == country_value) & (self.merge_compare_result['계정'] == row['Account'])].index.min()# 각 계정이 삽입된 최소 인덱스를 추출
            if pd.notna(match_idx):  # 일치하는 인덱스가 존재할 경우
                self.merge_compare_result.at[match_idx, 'NA'] = rec_product # 해당 인덱스의 na 컬럼에 추천 제품 데이터 삽입
        print("이부분 확인해야함 : ",self.merge_compare_result)

    def get_result(self):
        """
        비교 결과를 Excel 파일로 저장하고 포맷팅하는 함수
        
        - 국가별로 시트를 분리하여 저장
        - Excel 파일의 컬럼 너비와 셀 정렬 설정
        - 추천 제품 컬럼을 계정별로 병합
        - 각 계정의 마지막 행에 굵은 테두리 추가
        """
        # Excel 파일로 저장 (openpyxl 엔진 사용)
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            for country in self.country_code:
            # 조건에 맞는 데이터 추출
                filtered_df = self.merge_compare_result[self.merge_compare_result['국가'] == country]

                # 시트 이름으로 저장
                sheet_name = f'비교결과({country})'  # Excel 시트명은 31자 제한
                filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
        #sheet_name = "비교결과("+self.country_code+")"  # 주석 처리된 이전 방식

                #self.merge_compare_result.to_excel(self.output_path,header=True, index=False, sheet_name=sheet_name)  # 주석 처리된 이전 방식
        
        # Excel 파일 포맷팅
        wb = load_workbook(self.output_path)
        for country in self.country_code:        
            
            sheet_name = f'비교결과({country})'
            ws = wb[sheet_name]
            
            # 엑셀의 D컬럼위치와 E컬럼 위치 너비 조정
            for col_index in range(1, ws.max_column+1):
                col_letter = get_column_letter(col_index)
                if(col_letter=="D" or col_letter=="E"):  # 결과 상세와 NA 컬럼은 넓게
                    ws.column_dimensions[col_letter].width = 50
                else:  # 나머지 컬럼들은 기본 너비
                     ws.column_dimensions[col_letter].width = 20  
            
            # 각 셀별 텍스트 위치 조정
            for row in ws.iter_rows():
            
                for cell in row:
                    cell.alignment = Alignment(vertical='center',horizontal='center')  # 기본 정렬: 가운데
                    rec_col = row[3]  # 결과 상세 컬럼 (D열)
                    rec_col.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)  # 왼쪽 정렬, 자동 줄바꿈
                    rec_col = row[4]  # NA 컬럼 (E열)
                    rec_col.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)  # 위쪽 정렬, 자동 줄바꿈
            
            last_row =0
            start_row=1
            bottom_thick_border = Border(bottom=Side(border_style="thick", color="000000"))  # 굵은 아래 테두리
            
            # 추천제품 컬럼을 각 계정에 맞게 병합
            for num, value in self.df_abs_data.drop_duplicates(subset=['Account']).iterrows():
                
                target_account = value['Account']  # 대상 계정
                 
                for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    
                      # 헤더 제외하고 행 순회
                    if row[0] == target_account:  # 계정 컬럼이 첫 번째 열(A열)이라고 가정
                        
                        last_row = idx  # 해당 계정의 마지막 행 인덱스 업데이트
                     
                ws.merge_cells(start_row=start_row+1, start_column=5, end_row=last_row, end_column=5)  # NA 컬럼(E열) 병합
                
                # 반복 대상 행 (예: 병합된 마지막 행) -> 마지막행마다 아래 라인을 굵게 표시
                for col in range(1, 6):  # 1~5열
                    cell = ws.cell(row=last_row, column=col)
                    cell.border = bottom_thick_border  # 굵은 아래 테두리 적용
                start_row = last_row  # 다음 병합을 위한 시작 행 업데이트
        
            # 4. 저장
            ws.delete_cols(6)  # 6번째 컬럼(국가 컬럼) 삭제
        wb.save(self.output_path)  # 변경사항 저장


